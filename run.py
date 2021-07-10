from openpyxl import load_workbook
import json

wb = load_workbook(filename = '20210624.xlsx')
sheet = wb.active

master = {}
master['cities'] = []
data = master['cities']

for row in sheet.iter_rows(min_row=2, max_row=50, max_col=4, values_only=True):
    city = row[0]
    city = city.rstrip();

    district = row[1]
    district = district.rstrip();

    neighborhood = row[3]
    neighborhood = neighborhood.rstrip();
    
    foundCity = False
    foundDistrict = False
    foundNeighborhood = False
    
    for x in data:
        if x['name'] == city:
            for y in x['districts']:
                if y['name'] == district:
                    for z in y['neighborhoods']:
                        if z == neighborhood:
                            
                            foundNeighborhood = True

                    if not foundNeighborhood:
                        y['neighborhoods'].append(neighborhood)
                    
                    foundDistrict = True

            if not foundDistrict:
                x['districts'].append({
                    'name': district,
                    'neighborhoods': []
                })
            
            foundCity = True

    if not foundCity:
        data.append({
            'name': city,
            'districts': []
        })

with open('output.json', 'w') as outfile:
    json.dump(master, outfile)
